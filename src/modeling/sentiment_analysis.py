import os
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from transformers import AutoTokenizer, AutoModelForSequenceClassification, pipeline
import torch
import re
from collections import defaultdict
import glob
from tqdm import tqdm
import warnings
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from matplotlib.lines import Line2D
import warnings

warnings.filterwarnings('ignore')


DOCUMENTS_DIR = BASE_DIR
MODEL_DIR = os.path.join(BASE_DIR, "bert-base-japanese-v3-finetuned-sentiment")
DOCUMENT_RESULT_CSV = os.path.join(BASE_DIR, "document_sentiment_analysis_5class.csv")
SENTENCE_RESULT_EXCEL = os.path.join(BASE_DIR, "sentence_sentiment_analysis.xlsx")
VISUALIZATION_PATH = os.path.join(BASE_DIR, "document_sentiment_trends_5class.png")


def clean_text_for_excel(text):
    if not isinstance(text, str):
        return str(text)

    cleaned = []
    for char in text:
        ascii_val = ord(char)
        if 0 <= ascii_val <= 31 and ascii_val not in [9, 10, 13]:
            continue  
        cleaned.append(char)

    result = ''.join(cleaned)


    result = re.sub(r'[\u200B-\u200D\uFEFF]', '', result)


    if len(result) > 32767: 
        result = result[:32700] + "...[已截断]"

    return result


def split_japanese_sentences(text, max_length=450):

    if not text or not isinstance(text, str):
        return []
        
    sentence_delimiters = r'[。！？!?]|\n'
    sentences = re.split(sentence_delimiters, text)

   
    sentences = [clean_text_for_excel(s.strip()) for s in sentences if s.strip()]

    return sentences


def analyze_sentiment_with_bert_5class(text, sentiment_analyzer, max_length=512):

    if not text or len(text) < 3:
        return None, 0.0, 0.0  

    try:

        if len(text) > max_length:
            text = text[:max_length]


        result = sentiment_analyzer(text)


        if isinstance(result, list) and len(result) > 0:
            label = result[0]['label']
            confidence = result[0]['score']

            label_mapping = {
                'LABEL_0': {'chinese': '十分消极', 'value': 0},
                'LABEL_1': {'chinese': '比较消极', 'value': 1},
                'LABEL_2': {'chinese': '中性', 'value': 2},
                'LABEL_3': {'chinese': '比较积极', 'value': 3},
                'LABEL_4': {'chinese': '十分积极', 'value': 4}
            }


            if label in label_mapping:
                sentiment_info = label_mapping[label]
                return sentiment_info['chinese'], sentiment_info['value'], confidence
            else:

                alternative_mapping = {
                    '0': {'chinese': '十分消极', 'value': 0},
                    '1': {'chinese': '比较消极', 'value': 1},
                    '2': {'chinese': '中性', 'value': 2},
                    '3': {'chinese': '比较积极', 'value': 3},
                    '4': {'chinese': '十分积极', 'value': 4},
                    'negative': {'chinese': '十分消极', 'value': 0},
                    'slightly_negative': {'chinese': '比较消极', 'value': 1},
                    'neutral': {'chinese': '中性', 'value': 2},
                    'slightly_positive': {'chinese': '比较积极', 'value': 3},
                    'positive': {'chinese': '十分积极', 'value': 4}
                }

                if str(label) in alternative_mapping:
                    sentiment_info = alternative_mapping[str(label)]
                    return sentiment_info['chinese'], sentiment_info['value'], confidence
                else:
                    print(f"未知标签: {label}, 置信度: {confidence}")
                    return '未知', 2, confidence 

        return '未知', 2, 0.0

    except Exception as e:
        print(f"情感分析出错: {e}, 文本: {text[:50]}...")
        return None, 2, 0.0


def read_documents_from_folder(folder_path):

    documents = []


    txt_files = []
    for root, dirs, files in os.walk(folder_path):
        if MODEL_DIR in root:
            continue

        for file in files:
            if file.endswith('.txt'):
                file_path = os.path.join(root, file)
                txt_files.append(file_path)


    txt_files.sort()

    print(f"找到 {len(txt_files)} 个txt文件")

    for file_path in tqdm(txt_files, desc="读取文档"):
        try:
            encodings = ['utf-8', 'shift_jis', 'euc-jp', 'cp932', 'utf-16', 'latin-1']
            content = None

            for encoding in encodings:
                try:
                    with open(file_path, 'r', encoding=encoding) as f:
                        content = f.read()
                    break
                except UnicodeDecodeError:
                    continue

            if content is None:

                with open(file_path, 'rb') as f:
                    content = f.read().decode('utf-8', errors='ignore')


            content = clean_text_for_excel(content)


            filename = os.path.splitext(os.path.basename(file_path))[0]
            filename = clean_text_for_excel(filename)

            documents.append({
                'file_path': file_path,
                'filename': filename,
                'content': content,
                'length': len(content)
            })

        except Exception as e:
            print(f"读取文件 {file_path} 时出错: {e}")
            continue

    return documents


def analyze_document_sentiment_5class(document, sentiment_analyzer):

    try:
        content = document['content']

        sentences = split_japanese_sentences(content)

        if not sentences:
            return {
                'filename': document['filename'],
                'file_path': document['file_path'],
                'sentence_count': 0,
                'avg_sentiment_value': 2.0,  
                'avg_confidence': 0.0,
                'sentiment_distribution': {'十分消极': 0, '比较消极': 0, '中性': 0, '比较积极': 0, '十分积极': 0},
                'dominant_sentiment': '中性',
                'sentence_results': []
            }


        sentence_results = []
        sentiment_values = []
        confidences = []
        sentiment_counts = defaultdict(int)

        for sentence in sentences:
            if sentence and len(sentence) > 1:
                try:
                    sentiment_chinese, sentiment_value, confidence = analyze_sentiment_with_bert_5class(sentence,
                                                                                                        sentiment_analyzer)

                    if sentiment_chinese is not None:
                        cleaned_sentence = clean_text_for_excel(sentence)

                        sentence_results.append({
                            'sentence': cleaned_sentence,
                            'sentiment_chinese': clean_text_for_excel(sentiment_chinese),
                            'sentiment_value': sentiment_value,
                            'confidence': confidence
                        })

                        sentiment_values.append(sentiment_value)
                        confidences.append(confidence)
                        sentiment_counts[sentiment_chinese] += 1
                except Exception as e:
                    print(f"处理句子时出错: {e}, 句子: {sentence[:50]}...")
                    continue


        if sentiment_values:
            avg_sentiment_value = np.mean(sentiment_values)
            avg_confidence = np.mean(confidences)


            if sentiment_counts:
                dominant_sentiment = max(sentiment_counts, key=sentiment_counts.get)
            else:
                dominant_sentiment = '中性'
        else:
            avg_sentiment_value = 2.0
            avg_confidence = 0.0
            dominant_sentiment = '中性'

        return {
            'filename': clean_text_for_excel(document['filename']),
            'file_path': clean_text_for_excel(document['file_path']),
            'sentence_count': len(sentence_results),
            'avg_sentiment_value': avg_sentiment_value,
            'avg_confidence': avg_confidence,
            'sentiment_distribution': dict(sentiment_counts),
            'dominant_sentiment': clean_text_for_excel(dominant_sentiment),
            'sentence_results': sentence_results
        }

    except Exception as e:
        print(f"分析文档 {document['filename']} 时出错: {e}")
        return None


def process_documents_5class(sentiment_analyzer):

    results = []
    all_sentence_results = [] 

    print("正在读取文档...")
    documents = read_documents_from_folder(DOCUMENTS_DIR)

    if not documents:
        print("没有找到可分析的文档")
        return results, all_sentence_results

    print(f"开始分析 {len(documents)} 个文档...")


    for idx, document in enumerate(tqdm(documents, desc="处理文档")):
        result = analyze_document_sentiment_5class(document, sentiment_analyzer)
        if result:
            result['doc_index'] = idx + 1
            results.append(result)

            for sentence_result in result['sentence_results']:
                sentence_result['doc_index'] = idx + 1
                sentence_result['filename'] = result['filename']
                all_sentence_results.append(sentence_result)

    return results, all_sentence_results


def save_results_5class(document_results, sentence_results):

    if not document_results:
        print("没有文档结果可保存")
        return None


    doc_data = []
    for result in document_results:
        dist = result['sentiment_distribution']
        doc_data.append({
            '文档序号': result['doc_index'],
            '文件名': result['filename'],
            '句子数量': result['sentence_count'],
            '平均情感数值': result['avg_sentiment_value'],
            '平均信度': result['avg_confidence'],
            '主导情感': result['dominant_sentiment'],
            '十分消极数量': dist.get('十分消极', 0),
            '比较消极数量': dist.get('比较消极', 0),
            '中性数量': dist.get('中性', 0),
            '比较积极数量': dist.get('比较积极', 0),
            '十分积极数量': dist.get('十分积极', 0),
            '文件路径': result['file_path']
        })

    df_doc = pd.DataFrame(doc_data)
    df_doc = df_doc.sort_values('文档序号')
    df_doc.to_csv(DOCUMENT_RESULT_CSV, index=False, encoding='utf-8-sig')
    print(f"文档级别结果已保存到: {DOCUMENT_RESULT_CSV}")


    if sentence_results:
        print("正在保存句子级别结果到Excel...")

        wb = Workbook()
        ws = wb.active
        ws.title = "情感分析结果"

        headers = ['文档序号', '文件名', '日语句子', '情感极性', '情感数值', '极性信度']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # 设置标题样式
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")


        row = 2
        for sentence_result in sentence_results:
            try:
                ws.cell(row=row, column=1, value=sentence_result['doc_index'])
                ws.cell(row=row, column=2, value=clean_text_for_excel(str(sentence_result['filename'])))
                ws.cell(row=row, column=3, value=clean_text_for_excel(str(sentence_result['sentence'])))
                ws.cell(row=row, column=4, value=clean_text_for_excel(str(sentence_result['sentiment_chinese'])))
                ws.cell(row=row, column=5, value=sentence_result['sentiment_value'])
                ws.cell(row=row, column=6, value=sentence_result['confidence'])
                row += 1
            except Exception as e:
                print(f"写入Excel行 {row} 时出错: {e}")
                print(f"问题数据: {sentence_result.get('sentence', '')[:50]}")
                continue


        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


        try:
            wb.save(SENTENCE_RESULT_EXCEL)
            print(f"句子级别结果已保存到: {SENTENCE_RESULT_EXCEL}")
        except Exception as e:
            print(f"保存Excel文件失败: {e}")
            backup_path = SENTENCE_RESULT_EXCEL.replace('.xlsx', '_backup.csv')
            sentence_df = pd.DataFrame(sentence_results)
            sentence_df.to_csv(backup_path, index=False, encoding='utf-8-sig')
            print(f"已保存为CSV备份文件: {backup_path}")

    return df_doc


def create_sentiment_trend_visualization(df):

    if df is None or len(df) == 0:
        print("没有数据可生成可视化")
        return


    try:
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
    except:
        print("注意：可能无法正确显示中文字符")


    fig, ax1 = plt.subplots(figsize=(15, 8))


    x = df['文档序号']


    color1 = 'tab:blue'
    ax1.set_xlabel('文档序号', fontsize=12)
    ax1.set_ylabel('平均情感数值 (0-4)', color=color1, fontsize=12)
    line1 = ax1.plot(x, df['平均情感数值'], color=color1, linewidth=2, marker='o', markersize=4, label='平均情感数值')
    ax1.tick_params(axis='y', labelcolor=color1)
    ax1.grid(True, alpha=0.3, linestyle='--')
    ax1.set_ylim([0, 4])


    ax1.axhline(y=0, color='red', linestyle=':', alpha=0.5, label='十分消极')
    ax1.axhline(y=1, color='orange', linestyle=':', alpha=0.5, label='比较消极')
    ax1.axhline(y=2, color='gray', linestyle=':', alpha=0.5, label='中性')
    ax1.axhline(y=3, color='lightgreen', linestyle=':', alpha=0.5, label='比较积极')
    ax1.axhline(y=4, color='green', linestyle=':', alpha=0.5, label='十分积极')


    ax2 = ax1.twinx()
    color2 = 'tab:red'
    ax2.set_ylabel('平均信度 (0-1)', color=color2, fontsize=12)


    confidence_min = df['平均信度'].min()
    confidence_max = df['平均信度'].max()
    confidence_avg = df['平均信度'].mean()

    print(f"信度统计: 最小值={confidence_min:.4f}, 最大值={confidence_max:.4f}, 平均值={confidence_avg:.4f}")


    if confidence_max - confidence_min < 0.1:  

        confidence_range = max(0.05, confidence_max - confidence_min)
        confidence_mid = (confidence_min + confidence_max) / 2
        ax2.set_ylim([
            max(0, confidence_mid - confidence_range * 2),
            min(1, confidence_mid + confidence_range * 2)
        ])
    else:
        ax2.set_ylim([0, 1])


    scatter2 = ax2.scatter(x, df['平均信度'], color=color2, s=50, alpha=0.7,
                           label='平均信度', edgecolors='black')

    ax2.tick_params(axis='y', labelcolor=color2)


    legend_elements = [
        Line2D([0], [0], color=color1, lw=2, label='平均情感数值'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor=color2,
               markersize=10, label='平均信度', markeredgecolor='black')
    ]
    ax1.legend(handles=legend_elements, loc='upper left')

    plt.title('文档情感变化趋势图 (双纵轴)', fontsize=16, pad=20)
    plt.tight_layout()


    try:
        plt.savefig(VISUALIZATION_PATH, dpi=300, bbox_inches='tight',
                    metadata={'Creation Time': None})
    except:
        plt.savefig(VISUALIZATION_PATH, dpi=300, bbox_inches='tight')

    print(f"情感趋势图已保存到: {VISUALIZATION_PATH}")


    plt.show()


def create_separate_visualizations(df):

    if df is None or len(df) == 0:
        print("没有数据可生成可视化")
        return


    try:
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
    except:
        print("注意：可能无法正确显示中文字符")

    plt.figure(figsize=(15, 6))
    x = df['文档序号']


    plt.subplot(1, 2, 1)
    plt.plot(x, df['平均情感数值'], color='tab:blue', linewidth=2, marker='o', markersize=4)
    plt.xlabel('文档序号')
    plt.ylabel('平均情感数值 (0-4)')
    plt.title('文档情感数值变化趋势')
    plt.grid(True, alpha=0.3)
    plt.ylim([0, 4])


    plt.axhline(y=0, color='red', linestyle=':', alpha=0.5, label='十分消极')
    plt.axhline(y=1, color='orange', linestyle=':', alpha=0.5, label='比较消极')
    plt.axhline(y=2, color='gray', linestyle=':', alpha=0.5, label='中性')
    plt.axhline(y=3, color='lightgreen', linestyle=':', alpha=0.5, label='比较积极')
    plt.axhline(y=4, color='green', linestyle=':', alpha=0.5, label='十分积极')


    plt.subplot(1, 2, 2)
    plt.scatter(x, df['平均信度'], color='tab:red', s=50, alpha=0.7, edgecolors='black')
    plt.xlabel('文档序号')
    plt.ylabel('平均信度 (0-1)')
    plt.title('文档分析信度变化')
    plt.grid(True, alpha=0.3)


    confidence_range = df['平均信度'].max() - df['平均信度'].min()
    if confidence_range < 0.1:
        confidence_mid = (df['平均信度'].max() + df['平均信度'].min()) / 2
        plt.ylim([max(0, confidence_mid - 0.05), min(1, confidence_mid + 0.05)])
    else:
        plt.ylim([0, 1])

    plt.tight_layout()


    separate_path = os.path.join(BASE_DIR, "separate_sentiment_confidence.png")
    try:
        plt.savefig(separate_path, dpi=300, bbox_inches='tight',
                    metadata={'Creation Time': None})
    except:
        plt.savefig(separate_path, dpi=300, bbox_inches='tight')

    print(f"独立图表已保存到: {separate_path}")
    plt.show()


def create_boxplot_visualizations(df):

    if df is None or len(df) == 0:
        print("没有数据可生成可视化")
        return


    try:
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
    except:
        print("注意：可能无法正确显示中文字符")


    fig, axes = plt.subplots(1, 2, figsize=(14, 6))


    ax1 = axes[0]
    x = df['文档序号']
    ax1.plot(x, df['平均情感数值'], color='tab:blue', linewidth=2, marker='o', markersize=4)
    ax1.set_xlabel('文档序号')
    ax1.set_ylabel('平均情感数值 (0-4)')
    ax1.set_title('文档情感数值变化趋势')
    ax1.grid(True, alpha=0.3)
    ax1.set_ylim([0, 4])


    ax2 = axes[1]
    ax2.boxplot(df['平均信度'], vert=True, patch_artist=True,
                boxprops=dict(facecolor='lightcoral'),
                medianprops=dict(color='red', linewidth=2))
    ax2.set_ylabel('平均信度')
    ax2.set_title('文档平均信度分布')
    ax2.set_xticklabels(['所有文档'])
    ax2.grid(True, alpha=0.3, axis='y')


    ax2.scatter([1] * len(df), df['平均信度'], alpha=0.5, color='tab:red', s=20)

    plt.tight_layout()


    boxplot_path = os.path.join(BASE_DIR, "sentiment_boxplot.png")
    try:
        plt.savefig(boxplot_path, dpi=300, bbox_inches='tight',
                    metadata={'Creation Time': None})
    except:
        plt.savefig(boxplot_path, dpi=300, bbox_inches='tight')

    print(f"箱线图已保存到: {boxplot_path}")


    print("\n平均信度详细统计:")
    print(f"最小值: {df['平均信度'].min():.6f}")
    print(f"25%分位数: {df['平均信度'].quantile(0.25):.6f}")
    print(f"中位数: {df['平均信度'].median():.6f}")
    print(f"75%分位数: {df['平均信度'].quantile(0.75):.6f}")
    print(f"最大值: {df['平均信度'].max():.6f}")
    print(f"标准差: {df['平均信度'].std():.6f}")

    plt.show()


def create_additional_visualizations(df):


    try:
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
        plt.rcParams['axes.unicode_minus'] = False
    except:
        print("注意：可能无法正确显示中文字符")


    plt.figure(figsize=(12, 5))
    plt.subplot(1, 2, 1)
    plt.hist(df['平均情感数值'], bins=20, alpha=0.7, color='skyblue', edgecolor='black')
    plt.xlabel('平均情感数值')
    plt.ylabel('文档数量')
    plt.title('平均情感数值分布')
    plt.axvline(x=df['平均情感数值'].mean(), color='red', linestyle='--',
                label=f'均值: {df["平均情感数值"].mean():.3f}')
    plt.legend()
    plt.grid(True, alpha=0.3)


    plt.subplot(1, 2, 2)
    sentiment_counts = df['主导情感'].value_counts()
    colors = {'十分消极': 'darkred', '比较消极': 'lightcoral', '中性': 'lightgray',
              '比较积极': 'lightgreen', '十分积极': 'darkgreen'}
    color_list = [colors.get(s, 'blue') for s in sentiment_counts.index]

    wedges, texts, autotexts = plt.pie(sentiment_counts.values, labels=sentiment_counts.index,
                                       autopct='%1.1f%%', colors=color_list, startangle=90)
    plt.title('主导情感分布')

    plt.tight_layout()


    chart1_path = os.path.join(BASE_DIR, "sentiment_distribution_charts.png")
    try:
        plt.savefig(chart1_path, dpi=300, bbox_inches='tight',
                    metadata={'Creation Time': None})
    except:
        plt.savefig(chart1_path, dpi=300, bbox_inches='tight')


    plt.figure(figsize=(10, 6))
    plt.hist(df['平均信度'], bins=20, alpha=0.7, color='orange', edgecolor='black')
    plt.xlabel('平均信度')
    plt.ylabel('文档数量')
    plt.title('平均信度分布')
    plt.axvline(x=df['平均信度'].mean(), color='red', linestyle='--',
                label=f'均值: {df["平均信度"].mean():.3f}')
    plt.legend()
    plt.grid(True, alpha=0.3)
    plt.tight_layout()


    chart2_path = os.path.join(BASE_DIR, "confidence_distribution.png")
    try:
        plt.savefig(chart2_path, dpi=300, bbox_inches='tight',
                    metadata={'Creation Time': None})
    except:
        plt.savefig(chart2_path, dpi=300, bbox_inches='tight')

    plt.show()


def load_sentiment_analyzer():

    print("正在加载BERT日语情感分析模型...")

    try:

        if os.path.exists(MODEL_DIR):
            print("从本地目录加载模型...")


            device = 0 if torch.cuda.is_available() else -1


            try:

                sentiment_analyzer = pipeline(
                    "sentiment-analysis",
                    model=MODEL_DIR,
                    tokenizer=MODEL_DIR,
                    device=device
                )
            except:

                print("使用简化加载方式...")
                sentiment_analyzer = pipeline(
                    "sentiment-analysis",
                    model=MODEL_DIR,
                    device=-1  
                )

            print("模型加载成功!")


            test_result = sentiment_analyzer("テスト文です。")
            print(f"模型测试结果: {test_result}")
            print(f"结果类型: {type(test_result)}")
            if isinstance(test_result, list) and len(test_result) > 0:
                print(f"标签: {test_result[0].get('label')}")
                print(f"信度: {test_result[0].get('score')}")

            return sentiment_analyzer
        else:
            print(f"错误：本地模型目录 {MODEL_DIR} 不存在")
            print("请确保bert-base-japanese-v3-finetuned-sentiment文件夹存在")
            return None

    except Exception as e:
        print(f"模型加载失败: {e}")
        print("尝试从Hugging Face下载模型...")

        try:

            sentiment_analyzer = pipeline(
                "sentiment-analysis",
                model="Mizuiro-sakura/bert-base-japanese-v3-finetuned-sentiment",
                device=-1
            )
            print("模型下载成功!")
            return sentiment_analyzer
        except Exception as e2:
            print(f"下载也失败: {e2}")
            return None


def main():

    print("=" * 60)
    print("日语文档情感分析系统 (5个情感类别)")
    print("=" * 60)


    if not os.path.exists(DOCUMENTS_DIR):
        print(f"错误：文档文件夹 {DOCUMENTS_DIR} 不存在")
        return


    sentiment_analyzer = load_sentiment_analyzer()
    if sentiment_analyzer is None:
        print("模型加载失败，程序退出")
        return


    print("\n" + "=" * 60)
    print("开始进行文档级情感分析 (5个类别)...")
    print("=" * 60)

    document_results, sentence_results = process_documents_5class(sentiment_analyzer)

    if not document_results:
        print("没有分析到有效结果")
        return


    print("\n" + "=" * 60)
    print("正在保存分析结果...")
    print("=" * 60)

    df = save_results_5class(document_results, sentence_results)


    print("\n" + "=" * 60)
    print("正在生成可视化图表...")
    print("=" * 60)

    
    create_sentiment_trend_visualization(df)
    create_separate_visualizations(df)
    create_boxplot_visualizations(df)
    create_additional_visualizations(df)

    print("\n" + "=" * 60)
    print("分析完成！")
    print("=" * 60)
    print(f"文档级别结果: {DOCUMENT_RESULT_CSV}")
    print(f"句子级别结果: {SENTENCE_RESULT_EXCEL}")
    print(f"情感趋势图: {VISUALIZATION_PATH}")


if __name__ == "__main__":
    main()
